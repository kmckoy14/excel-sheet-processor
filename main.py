
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime

def process_excel_files(folder_path, output_folder='processed_files', processed_log_file='processed_files.txt'):
    folder_path = os.path.normpath(folder_path)
    output_folder = os.path.normpath(output_folder)
    os.makedirs(output_folder, exist_ok=True)
    
    try:
        with open(processed_log_file, 'r') as f:
            processed_files = set(f.read().splitlines())
    except FileNotFoundError:
        processed_files = set()
        print(f"Created new processed files log: {processed_log_file}")
    
    excel_files = [f for f in os.listdir(folder_path) 
                  if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    if not excel_files:
        print(f"No Excel files found in {folder_path}")
        return
    
    for file_name in excel_files:
        if file_name in processed_files:
            print(f"Skipping already processed file: {file_name}")
            continue
            
        file_path = os.path.join(folder_path, file_name)
        
        try:
            # Read the Excel file starting from row 10
            df = pd.read_excel(file_path, engine='openpyxl', skiprows=10)
            
            print(f"Processing file: {file_name}")
            
            # Forward fill the PO Number - this is always needed
            df['PO Number'] = df['PO Number'].fillna(method='ffill')
            
            # Create a flag to identify where Total is explicitly set vs forward-filled
            df['Total_Is_Original'] = df['Total'].notna()
            
            # Now fill the Total values
            df['Total'] = df['Total'].fillna(method='ffill')
            
            # Only keep values where Total is explicitly set OR Item Key is not None
            df_cleaned = df.copy()
            for idx in range(1, len(df)):
                # If this row has the same PO Number as the previous row
                # AND the Total value is the same (was filled)
                # AND this row doesn't have an explicit Total value
                if (df.iloc[idx]['PO Number'] == df.iloc[idx-1]['PO Number'] and
                    df.iloc[idx]['Total'] == df.iloc[idx-1]['Total'] and
                    not df.iloc[idx]['Total_Is_Original']):
                    # Clear the Total for this row since it's just a duplicate from forward-fill
                    df_cleaned.at[idx, 'Total'] = None
            
            # Drop rows where Item Key is NaN (these might be empty rows or headers)
            df_cleaned = df_cleaned.dropna(subset=['Item Key'])
            
            # Extract required columns
            extracted_data = df_cleaned[[
                'PO Number',
                'Item Key',
                'Item Desc',
                'Total'
            ]].copy()
            
            # Rename columns
            extracted_data = extracted_data.rename(columns={'Item Key': 'Key'})
            
            # Add Source File column
            extracted_data.insert(1, 'Source_File', file_name)
            
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            
            # Add headers
            headers = ['PO Number', 'Source File', 'Key', 'Item Desc', 'Total']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write the data
            for row_idx, row in extracted_data.iterrows():
                ws.cell(row=row_idx+2, column=1, value=row['PO Number'])
                ws.cell(row=row_idx+2, column=2, value=row['Source_File'])
                ws.cell(row=row_idx+2, column=3, value=row['Key'])
                ws.cell(row=row_idx+2, column=4, value=row['Item Desc'])
                ws.cell(row=row_idx+2, column=5, value=row['Total'])
            
            # Create output filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'processed_{os.path.splitext(file_name)[0]}_{timestamp}.xlsx'
            output_path = os.path.join(output_folder, output_filename)
            
            
            wb.save(output_path)
            
            # Add to processed files log
            with open(processed_log_file, 'a') as f:
                f.write(f"{file_name}\n")
            
            print(f"Successfully processed: {file_name}")
            print(f"Output saved as: {output_filename}")
            
        except Exception as e:
            print(f"Error processing {file_name}: {str(e)}")

if __name__ == "__main__":
    folder_path = "/mnt/c/Workspace/Test_Folder"
    process_excel_files(folder_path)